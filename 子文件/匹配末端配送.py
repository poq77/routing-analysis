from openpyxl import Workbook, load_workbook

import pandas as pd
import ast  # 用于将字符串安全地转换为列表

# 读取 .xlsx 文件中的特定 sheet

sheet_name = "线路简单"  # 将这里的 "Sheet1" 替换为你要读取的工作表名称

df = pd.read_excel("输入.xlsx", sheet_name=sheet_name)

# 将 path 字段的字符串转换回列表

df["path"] = df["path"].apply(ast.literal_eval)

# 将 DataFrame 转换回字典列表

routes = df.to_dict('records')


# 初始化最大长度为 0
max_length = 0
for route in routes:
    current_length = len(route["path"])
    if current_length > max_length:
        max_length = current_length

# 转运中心个数和线路个数
zyzx = max_length
xianlu = max_length - 1

# 创建 Excel 工作簿
wb = Workbook()
ws = wb.active

# 定义转运中心相关属性列表
attributes = ["派送类型", "处理时长（小时）", "开始时间", "开始日期", "完成时间", "完成日期", "折算：元/票", "元/公斤", "元/票"]

# 读取末端配送.xlsx 文件
try:
    end_delivery_wb = load_workbook('输入.xlsx')
    end_delivery_ws = end_delivery_wb['末端配送']
except FileNotFoundError:
    print("未找到 '末端配送.xlsx' 文件，请检查文件是否存在。")
    exit()

# 获取末端配送表中出发地和目的地所在列的索引
header_row = [cell.value for cell in end_delivery_ws[1]]
origin_col_index = header_row.index('出发地') if '出发地' in header_row else None
destination_col_index = header_row.index('目的地') if '目的地' in header_row else None
if origin_col_index is None or destination_col_index is None:
    print("末端配送表中未找到 '出发地' 或 '目的地' 列，请检查文件格式。")
    exit()

# 获取派送类型列的索引
start_col_index = header_row.index('派送类型') if '派送类型' in header_row else None
if start_col_index is None:
    print("末端配送表中未找到 '派送类型' 列，请检查文件格式。")
    exit()

# 写入表头
ws.append(["出发地", "目的地"] + attributes)

# 遍历 routes 数据
for route in routes:
    origin = route["path"][-2]
    destination = route["path"][-1]
    # 在末端配送表中查找匹配的行
    for row in end_delivery_ws.iter_rows(min_row=2, values_only=True):
        if row[origin_col_index] == origin and row[destination_col_index] == destination:
            # 提取匹配行中从派送类型列到元/票列的值
            values = [origin, destination] + list(row[start_col_index:])
            ws.append(values)
            break

# 保存新的 Excel 文件
wb.save('outt.xlsx')
print("数据已处理并保存到 'output.xlsx' 文件中。")