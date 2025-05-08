from openpyxl import Workbook, load_workbook

# 定义路线数据
import pandas as pd
import ast  # 用于将字符串安全地转换为列表

# 读取 .xlsx 文件中的特定 sheet

sheet_name = "线路简单"  # 将这里的 "Sheet1" 替换为你要读取的工作表名称

df = pd.read_excel("输入.xlsx", sheet_name=sheet_name)

# 将 path 字段的字符串转换回列表

df["path"] = df["path"].apply(ast.literal_eval)

# 将 DataFrame 转换回字典列表

routes = df.to_dict('records')


# 找出最大路径长度
max_path_length = max(len(route["path"]) for route in routes)

# 计算转运中心个数和线路个数
num_transfer_centers = max_path_length
num_routes = max_path_length - 1

# 创建 Excel 工作簿
wb = Workbook()
ws = wb.active

# 定义转运中心相关属性列表
transfer_attributes = ["转运中心", "等待作业（小时）", "作业时长（小时）", "集货等待（小时）", "干线到达", "到达日期", "开始作业", "完成作业", "干线发出", "发出日期", "元/公斤", "元/票", "清关", "备注"]
# 处理转运中心属性，添加序号
transfer_header = [f"{attr} {i}" for i in range(1, num_transfer_centers + 1) for attr in transfer_attributes]

# 定义线路相关属性列表
route_attributes = ["线路", "出发地", "目的地", "运输时长（小时）", "距离（公里）", "发运时间（当地）", "发运日期", "到达时间（当地）", "到达日期", "元/公斤", "元/票", "清关", "配送"]
# 处理线路属性，添加序号
route_header = [f"{attr} {i}" for i in range(1, num_routes + 1) for attr in route_attributes]

# 构建表头
header = []
for i in range(max_path_length):
    start_transfer = i * len(transfer_attributes)
    end_transfer = start_transfer + len(transfer_attributes)
    header.extend(transfer_header[start_transfer:end_transfer])

    start_route = i * len(route_attributes)
    end_route = start_route + len(route_attributes)
    if start_route < len(route_header):
        header.extend(route_header[start_route:end_route])

# 写入表头
ws.append(header)

try:
    # 打开输入文件
    workbook = load_workbook('输入.xlsx')
    sheet_transfer = workbook['转运中心']
    sheet_route = workbook['线路详细']

    # 假设转运中心属性在第一列
    transfer_center_col_index = 1
    # 假设出发地属性在第二列
    origin_col_index = 2
    # 假设目的地属性在第三列
    destination_col_index = 3

    # 循环处理每条路线
    for route_idx, route in enumerate(routes, start=2):  # 从第二行开始写入数据
        # 处理转运中心数据
        transfer_data = []
        print(f"处理路线: {route['name']} - 转运中心")
        for transfer_center in route['path']:
            for row in sheet_transfer.iter_rows(min_row=2, values_only=True):
                if row[transfer_center_col_index - 1] == transfer_center:
                    transfer_data.extend(row)
                    break

        # 处理线路数据
        route_data = []
        print(f"处理路线: {route['name']} - 线路")
        for i in range(len(route['path']) - 1):
            origin = route['path'][i]
            destination = route['path'][i + 1]
            for row in sheet_route.iter_rows(min_row=2, values_only=True):
                if row[origin_col_index - 1] == origin and row[destination_col_index - 1] == destination:
                    route_data.extend(row)
                    break

        # 组合转运中心和线路数据
        combined_data = []
        for i in range(max_path_length):
            start_transfer = i * len(transfer_attributes)
            end_transfer = start_transfer + len(transfer_attributes)
            combined_data.extend(transfer_data[start_transfer:end_transfer])

            start_route = i * len(route_attributes)
            end_route = start_route + len(route_attributes)
            if start_route < len(route_data):
                combined_data.extend(route_data[start_route:end_route])

        # 写入 Excel
        for col_idx, value in enumerate(combined_data, start=1):
            ws.cell(row=route_idx, column=col_idx, value=value)

    # 保存工作簿
    wb.save("output.xlsx")
    print("数据已保存到 output.xlsx")
except FileNotFoundError:
    print("未找到输入文件 '输入.xlsx'，请检查文件路径和文件名。")
except KeyError as e:
    print(f"输入文件中未找到指定的工作表: {e}")
except Exception as e:
    print(f"发生未知错误: {e}")