from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

input_files = ["delivery_info.xlsx", "ut.xlsx", "output.xlsx", "outt.xlsx"]

merged_wb = Workbook()
merged_ws = merged_wb.active

current_col = 1

for file in input_files:
    try:
        wb = load_workbook(file)
        ws = wb.active
        max_rows = ws.max_row
        max_cols = ws.max_column

        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                value = ws.cell(row=row, column=col).value
                merged_ws.cell(row=row, column=current_col + col - 1, value=value)

        current_col += max_cols

    except FileNotFoundError:
        print(f"⚠️ 未找到文件: {file}，跳过。")

header_fill = PatternFill(start_color="4D148C", end_color="4D148C", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, name="微软雅黑")

for col in range(1, merged_ws.max_column + 1):
    cell = merged_ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

merged_wb.save("中乌路由.xlsx")
print("✅ 合并文件已保存为 中乌路由.xlsx")
