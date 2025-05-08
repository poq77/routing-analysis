from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# 定义需要合并的 Excel 文件列表
input_files = ["delivery_info.xlsx", "ut.xlsx", "output.xlsx", "outt.xlsx"]

# 创建一个新的工作簿
merged_wb = Workbook()
merged_ws = merged_wb.active

# 记录当前要写入的起始列
current_col = 1

# 遍历每个输入文件
for file in input_files:
    try:
        # 加载当前文件
        wb = load_workbook(file)
        # 获取第一个工作表
        ws = wb.active

        # 获取输入文件工作表的最大行数和列数
        max_rows = ws.max_row
        max_cols = ws.max_column

        # 遍历输入文件工作表的每一行和每一列
        for row in range(1, max_rows + 1):
            for col in range(1, max_cols + 1):
                # 获取输入文件单元格的值
                cell_value = ws.cell(row=row, column=col).value
                # 将值写入合并工作表对应的位置
                merged_ws.cell(row=row, column=current_col + col - 1, value=cell_value)

        # 更新下一个文件数据的起始列
        current_col += max_cols

    except FileNotFoundError:
        print(f"未找到文件: {file}，跳过该文件。")

# 设置首行样式（背景色、字体颜色、加粗）
header_fill = PatternFill(start_color="4D148C", end_color="4D148C", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, name="微软雅黑")

for col in range(1, merged_ws.max_column + 1):
    cell = merged_ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font

# 保存合并后的文件
merged_wb.save("中乌路由.xlsx")
print("文件合并完成，合并后的文件名为 中乌路由.xlsx。")