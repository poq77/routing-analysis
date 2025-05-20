from openpyxl import Workbook, load_workbook
import pandas as pd
import ast

sheet_name = "线路简单"
df = pd.read_excel("current_input.xlsx", sheet_name=sheet_name)
df["path"] = df["path"].apply(ast.literal_eval)
routes = df.to_dict('records')

max_path_length = max(len(route["path"]) for route in routes)
num_transfer_centers = max_path_length
num_routes = max_path_length - 1

wb = Workbook()
ws = wb.active

transfer_attributes = ["转运中心", "等待作业（小时）", "作业时长（小时）", "集货等待（小时）", "干线到达", "到达日期", "开始作业", "完成作业", "干线发出", "发出日期", "元/公斤", "元/票", "清关", "备注"]
transfer_header = [f"{attr} {i}" for i in range(1, num_transfer_centers + 1) for attr in transfer_attributes]

route_attributes = ["线路", "出发地", "目的地", "运输时长（小时）", "距离（公里）", "发运时间（当地）", "发运日期", "到达时间（当地）", "到达日期", "元/公斤", "元/票", "清关", "配送"]
route_header = [f"{attr} {i}" for i in range(1, num_routes + 1) for attr in route_attributes]

header = []
for i in range(max_path_length):
    header.extend(transfer_header[i * len(transfer_attributes):(i + 1) * len(transfer_attributes)])
    if i < num_routes:
        header.extend(route_header[i * len(route_attributes):(i + 1) * len(route_attributes)])
ws.append(header)

try:
    workbook = load_workbook("current_input.xlsx")
    sheet_transfer = workbook['转运中心']
    sheet_route = workbook['线路详细']

    for route_idx, route in enumerate(routes, start=2):
        transfer_data = []
        for transfer_center in route['path']:
            for row in sheet_transfer.iter_rows(min_row=2, values_only=True):
                if row[0] == transfer_center:
                    transfer_data.extend(row)
                    break

        route_data = []
        for i in range(len(route['path']) - 1):
            origin = route['path'][i]
            destination = route['path'][i + 1]
            for row in sheet_route.iter_rows(min_row=2, values_only=True):
                if row[1] == origin and row[2] == destination:
                    route_data.extend(row)
                    break

        combined_data = []
        for i in range(max_path_length):
            combined_data.extend(transfer_data[i * len(transfer_attributes):(i + 1) * len(transfer_attributes)])
            if i < num_routes:
                combined_data.extend(route_data[i * len(route_attributes):(i + 1) * len(route_attributes)])
        for col_idx, value in enumerate(combined_data, start=1):
            ws.cell(row=route_idx, column=col_idx, value=value)

    wb.save("output.xlsx")
    print("✅ 数据已保存到 output.xlsx")
except Exception as e:
    print(f"❌ 错误：{e}")
