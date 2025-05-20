from openpyxl import Workbook, load_workbook
import pandas as pd
import ast

df = pd.read_excel("current_input.xlsx", sheet_name="线路简单")
df["path"] = df["path"].apply(ast.literal_eval)
routes = df.to_dict('records')

zyzx = max(len(route["path"]) for route in routes)
xianlu = zyzx - 1

wb = Workbook()
ws = wb.active

attributes = ["派送类型", "处理时长（小时）", "开始时间", "开始日期", "完成时间", "完成日期", "折算：元/票", "元/公斤", "元/票"]

try:
    end_delivery_wb = load_workbook("current_input.xlsx")
    end_delivery_ws = end_delivery_wb['末端配送']
except FileNotFoundError:
    print("❌ 未找到 current_input.xlsx 文件")
    exit()

header_row = [cell.value for cell in end_delivery_ws[1]]
origin_col_index = header_row.index('出发地')
destination_col_index = header_row.index('目的地')
start_col_index = header_row.index('派送类型')

ws.append(["出发地", "目的地"] + attributes)

for route in routes:
    origin = route["path"][-2]
    destination = route["path"][-1]
    for row in end_delivery_ws.iter_rows(min_row=2, values_only=True):
        if row[origin_col_index] == origin and row[destination_col_index] == destination:
            values = [origin, destination] + list(row[start_col_index:])
            ws.append(values)
            break

wb.save('outt.xlsx')
print("✅ 已保存文件 outt.xlsx")
