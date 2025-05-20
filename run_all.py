import os

python_file_path = r'D:\VsCodeProjects\routing_analysis\scripts\generate_route_summary.py'
return_code = os.system(f'python {python_file_path}')


python_file_path = r'D:\VsCodeProjects\routing_analysis\scripts\generate_pickup.py'
return_code = os.system(f'python {python_file_path}')


python_file_path = r'D:\VsCodeProjects\routing_analysis\scripts\generate_transfer_and_line.py'
return_code = os.system(f'python {python_file_path}')


python_file_path = r'D:\VsCodeProjects\routing_analysis\scripts\match_last_mile.py'
return_code = os.system(f'python {python_file_path}')


python_file_path = r'D:\VsCodeProjects\routing_analysis\scripts\merge_outputs.py'
return_code = os.system(f'python {python_file_path}')

import openpyxl
import re

# 加载 Excel 文件
file_path = "D:\\VsCodeProjects\\routing_analysis\\中乌路由.xlsx"  # 请替换为你的文件路径
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # 选择活动工作表

# 读取 C1 单元格内容并提取 dp 数值
c1_value = ws["C1"].value
match = re.search(r"\d+", str(c1_value))
dp = int(match.group()) if match else 1  # 如果没有找到数字，默认 dp=1

# 获取最大数据行数（非空行）
max_row = ws.max_row

# 定义通用公式（不包含 U 列）
formulas = {
    "F": "=BK{row}+CL{row}+DM{row}+EN{row}",
    # "G" 在下面单独处理
    "H": "=FR{row}",
    "I": "=SUM(W{row}:AI{row})",
    "J": "=R{row}+T{row}+U{row}",
    "K": "=S{row}+V{row}",
    "L": f"=J{{row}}+K{{row}}*{dp}",
    "M": f"=R{{row}}*{dp}+S{{row}}",
    "N": f"=T{{row}}*{dp}",
    "O": "=MAX(EF{row},EG{row})",
    "P": f"=ES{{row}}*{dp}+ET{{row}}",
    "Q": "=FS{row}",
    "R": "=AQ{row}+BC{row}+BP{row}+CD{row}+CQ{row}+DE{row}",
    "S": "=AR{row}+BD{row}+BQ{row}+CE{row}+CR{row}+DF{row}",
    "T": "=DR{row}",
    # "U" 在下面单独处理
    "V": "=O{row}+ET{row}+FH{row}+FS{row}",
    "W": "=AL{row}",
    "X": "=BJ{row}+CK{row}",
    "Y": "=AT{row}+BU{row}+CV{row}",
    "Z": "=AU{row}+BV{row}+CW{row}",
    "AA": "=AV{row}+BW{row}+CX{row}",
    "AB": '="T+"&INT(SUM(W{row}:AA{row})/24)',
    "AC": "=DL{row}",
    "AD": '="T+"&INT(SUM(Y{row}:AC{row})/24)',
    "AE": "=EM{row}",
    "AF": "=DW{row}",
    "AG": "=DX{row}",
    "AH": "=DY{row}",
    "AI": "=FN{row}",
    "AJ": '="T+"&INT(SUM(AE{row}:AI{row})/24)',
}

# 遍历数据行（从第二行开始）
for row in range(2, max_row + 1):
    # 读取 A 列和 B 列值
    a_value = str(ws[f"A{row}"].value).strip() if ws[f"A{row}"].value else ""
    b_value = str(ws[f"B{row}"].value).strip() if ws[f"B{row}"].value else ""

    # 处理 U 列的动态公式
    if b_value == "线路2":
        u_formula = f"=ES{row}+FG{row}+FT{row}"
    else:
        u_formula = f"=ES{row}+FG{row}"
    ws[f"U{row}"] = u_formula  # 写入 U 列公式

    # 处理 G 列的动态值
    if b_value == "线路3" and "店配" in a_value:
        g_value = "T+28"
    else:
        g_value = "T+26"
    ws[f"G{row}"] = g_value  # 写入 G 列内容

    # 写入其他列的公式
    for col, formula in formulas.items():
        ws[f"{col}{row}"] = formula.format(row=row)

# 保存 Excel 文件
output_file = "D:\\VsCodeProjects\\routing_analysis\\routing_analysis.xlsx"
wb.save(output_file)

print(f"Excel 文件已更新，并保存为 '{output_file}'，G列和U列已动态填充，dp={dp}")
